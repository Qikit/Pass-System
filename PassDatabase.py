import random
import openpyxl
import qrcode

bd = "classroom.xlsx"
book = openpyxl.load_workbook(bd, data_only=True)
sheet = book.active
aut = int(sheet.max_row) + 1
auti = int(sheet.max_row)
id = "Пропускная система Гимназии\nИдентификационный номер ученика: "

print("Добро пожаловать в систему пропусков!")
print("Напишите, пожалуйста, какой класс вы добавляете?")
classroom = input("Класс: ")
lit = input("Литер класса: ")
print("Отлично! Какое количество учеников вы добавляете?")
count = int(input("Количество учеников: "))
for i in range(count):
    sheet['A' + str(aut + i)] = int(classroom)
    sheet['B' + str(aut + i)] = lit
    sheet['C' + str(aut + i)] = input("ФИО ученика: ")
    sheet['D' + str(aut + i)] = int(auti) + i
    sheet['E' + str(aut + i)] = int(str(classroom) + str((random.randint(1000, 9999))))
    print("Отлично! Добавлено:", str(sheet['A' + str(aut + i)].value) + str(sheet['B' + str(aut + i)].value),
          sheet['C' + str(aut + i)].value,
          sheet['D' + str(aut + i)].value, sheet['E' + str(aut + i)].value)
    book.save(bd)
    book.close()
    # text = id + str(sheet['D' + str(aut + i)].value) + "\n" + str(sheet['E' + str(aut + i)].value) + "\n" + str(
    #     sheet['C' + str(aut + i)].value) + " " + str(classroom) + str(lit)
    # path_to_download = Path().joinpath("", "logo.jpeg")
    # path_to_save = Path().joinpath("", "PassQR/" + str(classroom) + str(lit) + str(
    #     sheet['C' + str(aut + i)].value) + ".png")
    #
    # gen_qr_code(text, path_to_download, path_to_save)

    img = qrcode.make(
        id + str(sheet['D' + str(aut + i)].value) + "\n" + str(sheet['E' + str(aut + i)].value) + "\n" + str(
            sheet['C' + str(aut + i)].value) + " " + str(classroom) + str(lit))
    type(img)
    img.save("PassQR/" + str(classroom) + str(lit) + str(
        sheet['C' + str(aut + i)].value) + ".png")

print("В каталоге PassQr создан QR код.")

input("\nНажмите Enter для выхода")
